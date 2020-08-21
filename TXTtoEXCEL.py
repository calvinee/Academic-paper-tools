from openpyxl import Workbook,load_workbook

book_name_xlsx = r'D:\projects\临时操作文档\数据1.xlsx' #文件路径，把文档路径复制过来即可
wb = Workbook()
wb.save(book_name_xlsx)


#打开Excel
wb=load_workbook(book_name_xlsx)

#创建工作簿，导入的新数据会存在当前excel文件下新建一个‘s’的sheet里
sheet=wb.create_sheet('sheet1')

aa=[]
f = open(r'D:\projects\临时操作文档\IM.txt',encoding='utf-8')	#将从文献上的表格数据贴到txt文档中，将txt文档路径复制到此，encoding='utf-8'——编码格式
for line in f.readlines():
    data = line.split('\n\t')
    for str in data:
        sub_str = str.split(' ')	#每个数据间是按什么划分的，我的是两个空格符
        aa.append(sub_str)
        
for i in range(len(aa)):
    sheet.append(aa[i])
        #保存文件
    wb.save(book_name_xlsx)
    
print("题目写入数据成功！")
